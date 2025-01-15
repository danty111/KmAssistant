import os
import re

import openpyxl
from xmindparser import xmind_to_dict
import pandas as pd

from common.common_file import Common


class getVrs():
    """
    处理xmind转换成excel
    """
    def __init__(self):
        # 写入表头
        self.headers = {'test_suite': '测试用例集','labels': '用例标签', "module": '模块','requirement': '需求', 'priority': '优先级','title': '标题', 'description': '描述（前置条件）',
                   'step_id':'步骤ID','step': '步骤', 'test_data': '测试数据', 'expected_result': '期望结果',
                   'assignee': '经办人', 'reporter': '报告人'}
        self.makers_convert = {'1':'Highest','2':'High','3':'Medium','4':'low','5':'lowest'}
        self.case_list = {}
        self.case_name = ''
        self.flattened_list = []
        self.temporarily_list = []

    def convert_xmind_to_excel(self,xmind_file, excel_file):
        try:
            # 解析 XMind 文件
            xmind_data = xmind_to_dict(xmind_file)
        except:
            raise Exception("xmind文件损坏")
        # 创建 Excel 工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        for col, header in enumerate(self.headers.items(), start=1):
            sheet.cell(row=1, column=col, value=header[1])

        self.headers = {key: '' for key in self.headers}
        # 写入测试用例数据
        row = 2  # 从第二行开始写入数据
        self.process_topics(xmind_data[0]['topic'], sheet, row)
        # 保存 Excel 文件
        workbook.save(excel_file)

        # 读取 XLSX 文件
        df = pd.read_excel(excel_file)
        # 将 DataFrame 对象保存为 CSV 文件
        file_csv = excel_file.replace(".xlsx",".csv")
        df.to_csv(file_csv, index=False, encoding='utf-8')
        if os.path.exists(excel_file):
            os.rename(excel_file,excel_file.replace(".xlsx","_excel.xlsx"))

    def get_case_name(self, data, case_name=''):
        if isinstance(data, list):
            for item in data:
                yield from self.get_case_name(item, case_name)
        elif isinstance(data, dict):
            case_name = case_name + "/" + data['title']
            if 'makers' in data:
                remaining_data = data['topics'] if 'topics' in data else None
                yield case_name, data['makers'], remaining_data
            if 'topics' in data and isinstance(data['topics'], list):
                yield from self.get_case_name(data['topics'], case_name)
        else:
            raise ValueError("Invalid data format")

    def flatten_data(self,data):
        for item in data:
            if isinstance(item, dict) and 'title' in item:
                self.temporarily_list.append(item['title'])
            if 'topics' in item and isinstance(item['topics'], list):
                self.flatten_data(item['topics'])
            else:
                # 由于不一定存在描述，所以此处用倒叙
                self.temporarily_list.reverse()
                self.flattened_list.append(self.temporarily_list)
                self.temporarily_list = []
        return self.flattened_list

    def process_topics(self,topics, sheet,row):
        # 用例集
        test_suite = topics['title']
        topics = topics['topics']
        for topic in topics:

            # 用例列表
            case_list = topic['topics'][0]['topics']
            # 过滤掉需求标题下对应多条用例
            # case = self.get_case_name(case_list)
            # case_name, makers, case_list = next(case)
            # for case in case_list:
            my_generator = self.get_case_name(case_list)


            while True:
                try:
                    case_name, makers, remaining_data = next(my_generator)
                    # 用例集
                    self.headers['test_suite'] = test_suite
                    # 标题
                    self.headers["title"] = case_name
                    # 用例等级
                    try:
                        case_leverl = re.search(r'\d+', makers[0]).group()
                        self.headers["priority"] = self.makers_convert[case_leverl]
                    except Exception as e:
                        raise Exception(f"请检查{self.headers["title"]}的用例优先级，报错：",e)
                    # 模块
                    self.headers["module"] = topic['title']
                    # 需求
                    self.headers['requirement'] = topic['topics'][0]['title']

                    sheet.cell(row=row, column=self.get_key_index(self.headers, "test_suite"),
                               value=self.headers['test_suite'])
                    sheet.cell(row=row, column=self.get_key_index(self.headers, "module"),
                               value=self.headers["module"])

                    sheet.cell(row=row, column=self.get_key_index(self.headers, "requirement"),
                               value=self.get_demand_id(self.headers['requirement']))
                    # case_value = get_values(case_list[0])
                    case_value = self.flatten_data(remaining_data)
                    sheet.cell(row=row, column=self.get_key_index(self.headers, "priority"),
                               value=self.headers["priority"])

                    sheet.cell(row=row, column=self.get_key_index(self.headers, "title"),
                               value=self.headers["title"])
                    # #报告人

                    sheet.cell(row=row, column=self.get_key_index(self.headers, "reporter"),
                               value='博仔')


                    step_id = 0
                    for i in case_value:
                        step_id += 1
                        if len(i) == 3:
                            # 描述（前置条件）
                            self.headers['description'] = i[2]
                        # 步骤ID
                        self.headers['step_id'] = step_id
                        # 步骤
                        self.headers['step'] = i[1]
                        # 期望结果
                        self.headers['expected_result'] = i[0]
                        sheet.cell(row=row, column=self.get_key_index(self.headers, "step_id"), value=step_id)
                        sheet.cell(row=row, column=self.get_key_index(self.headers, "step"), value=self.headers['step'])
                        sheet.cell(row=row, column=self.get_key_index(self.headers, "description"),
                                   value=self.headers["description"])
                        sheet.cell(row=row, column=self.get_key_index(self.headers, "expected_result"),
                                   value=self.headers['expected_result'])
                        self.flattened_list = []
                        self.headers = {key: '' for key in self.headers}

                        row += 1
                except StopIteration:
                    break
                except Exception as e :
                    raise Exception("缺少数据",self.headers["title"],e)

    def get_demand_id(self,text):

        """
        传入需求名字，需再 中带入需求id
        :return:
        """
        pattern = r"[\[【](.*?)[\]】]"
        match = re.search(pattern, text)
        if match:
            card_number = match.group(1)
            return card_number
        raise Exception("id需要以【】或[]包裹id")


    def get_key_index(self,dictionary, target_key):
        for index, key in enumerate(dictionary):
            if key == target_key:
                return index + 1
        return -1

    def get_priority(self,priority):
        # 根据 XMind 中的优先级映射为 Jira 中的优先级
        priority_map = {
            '1': '最高',
            '2': '高',
            '3': '中',
            '4': '低',
            '5': '最低'
        }
        return priority_map.get(priority, '')

    def get_labels(labels):
        # 将 XMind 中的标签转换为 Jira 的标签格式
        return ','.join(labels)
